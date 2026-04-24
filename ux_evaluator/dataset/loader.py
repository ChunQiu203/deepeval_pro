import os
import json
import csv
import ast  # 用于安全地将字符串解析为列表/字典
from typing import List, Dict, Any, Optional


class TestCase:
    """标准化的测试用例结构，统一全项目的输入格式，兼容DeepEval框架"""
    def __init__(
        self,
        input: str,
        actual_output: str,
        expected_output: Optional[str] = None,
        context: Optional[List[str]] = None,
        retrieval_context: Optional[List[str]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ):
        """
        Args:
            input: 用户的输入/问题
            actual_output: LLM的实际输出
            expected_output: 期望的输出（可选）
            context: 对话上下文（可选）
            retrieval_context: RAG检索到的上下文（可选）
            metadata: 额外的元数据（可选）
        """
        self.input = input
        self.actual_output = actual_output
        self.expected_output = expected_output
        self.context = context or []
        self.retrieval_context = retrieval_context or []
        self.metadata = metadata or {}
    
    def to_dict(self) -> Dict[str, Any]:
        """转换为字典格式，方便序列化"""
        return {
            "input": self.input,
            "actual_output": self.actual_output,
            "expected_output": self.expected_output,
            "context": self.context,
            "retrieval_context": self.retrieval_context,
            "metadata": self.metadata
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> "TestCase":
        """从字典创建测试用例"""
        return cls(
            input=data.get("input", ""),
            actual_output=data.get("actual_output", ""),
            expected_output=data.get("expected_output"),
            context=data.get("context"),
            retrieval_context=data.get("retrieval_context"),
            metadata=data.get("metadata")
        )


class DatasetLoader:
    """数据集加载器，支持CSV/JSON格式导入，自动清洗与标准化数据"""
    
    def __init__(self, clean_data: bool = True):
        """
        Args:
            clean_data: 是否自动清洗数据（过滤空样本、标准化字段）
        """
        self.clean_data = clean_data
    
    def load_from_file(self, file_path: str) -> List[TestCase]:
        """
        从文件加载数据集，自动识别文件格式
        
        Args:
            file_path: 数据文件路径，支持.csv/.json/.xlsx
        
        Returns:
            标准化后的测试用例列表
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"数据文件不存在: {file_path}")
        
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            return self.load_from_csv(file_path)
        elif ext == ".json":
            return self.load_from_json(file_path)
        elif ext == ".xlsx":  # 仅支持 .xlsx
            return self.load_from_excel(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}，目前仅支持CSV、JSON和XLSX")

    def load_from_csv(self, file_path: str) -> List[TestCase]:
        """从CSV文件加载数据集"""
        raw_data = []
        # 使用 utf-8-sig 编码可以完美兼容带有 BOM 头（Excel保存）的 CSV 文件
        with open(file_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                raw_data.append(dict(row))

        return self._process_raw_data(raw_data)
    
    def load_from_json(self, file_path: str) -> List[TestCase]:
        """从JSON文件加载数据集，支持单行JSON或JSON数组"""
        with open(file_path, "r", encoding="utf-8") as f:
            try:
                # 尝试加载JSON数组
                raw_data = json.load(f)
                if not isinstance(raw_data, list):
                    raw_data = [raw_data]
            except json.JSONDecodeError:
                # 尝试加载JSONL（每行一个JSON）
                f.seek(0)
                raw_data = []
                for line in f:
                    line = line.strip()
                    if line:
                        raw_data.append(json.loads(line))
        
        return self._process_raw_data(raw_data)

    def load_from_excel(self, file_path: str) -> List[TestCase]:
        """从 Excel (.xlsx) 文件加载数据集"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("处理 Excel 文件需要安装 openpyxl 库，请运行: pip install openpyxl")

        try:
            # data_only=True 确保读取的是公式计算后的值，而不是公式本身
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            raw_data = []
            headers = []

            # iter_rows(values_only=True) 直接返回单元格的值，性能更好
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    # 第一行为表头，如果有空表头则自动生成一个占位符
                    headers = [str(cell).strip() if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
                else:
                    # 将数据行与表头组装成字典
                    row_dict = {}
                    for j, cell in enumerate(row):
                        # 处理空单元格，统一转换为空字符串，避免影响后续的 clean_data 判断
                        val = "" if cell is None else str(cell).strip()
                        # 防止由于数据列比表头列多导致的索引越界
                        if j < len(headers):
                            row_dict[headers[j]] = val

                    raw_data.append(row_dict)

            return self._process_raw_data(raw_data)
        except Exception as e:
            raise RuntimeError(f"加载 Excel 文件失败: {str(e)}")

    def _process_raw_data(self, raw_data: List[Dict[str, Any]]) -> List[TestCase]:
        """处理原始数据，清洗并标准化为TestCase"""

        def _parse_string_to_list(val: Any) -> Optional[List[str]]:
            """内部辅助函数：将 Excel 中的字符串列表智能转换为真正的 Python 列表"""
            if not val:
                return None
            if isinstance(val, list):
                return val
            if isinstance(val, str):
                val = val.strip()
                # 如果长得像列表 "['a', 'b']"
                if val.startswith('[') and val.endswith(']'):
                    try:
                        parsed = ast.literal_eval(val)
                        if isinstance(parsed, list):
                            return [str(i) for i in parsed]
                    except (ValueError, SyntaxError):
                        pass
                # 如果不是列表格式，就把它当成单个元素的列表
                return [val]
            return None

        def _parse_string_to_dict(val: Any) -> Dict[str, Any]:
            """内部辅助函数：将 Excel 中的字符串字典转换为真正的 Python 字典"""
            if isinstance(val, dict):
                return val
            if isinstance(val, str):
                val = val.strip()
                if val.startswith('{') and val.endswith('}'):
                    try:
                        parsed = ast.literal_eval(val)
                        if isinstance(parsed, dict):
                            return parsed
                    except (ValueError, SyntaxError):
                        pass
            return {}

        test_cases = []
        for item in raw_data:
            # 自动适配不同的字段名
            input_val = item.get("input", item.get("question", item.get("query", "")))
            output_val = item.get("actual_output", item.get("output", item.get("answer", "")))

            # 清洗数据：过滤空样本
            if self.clean_data:
                if not input_val or not output_val:
                    # 跳过空输入或空输出的样本
                    continue

            # 智能解析 retrieval_context 和 metadata
            retrieval_val = item.get("retrieval_context", item.get("知识库片段"))
            context_val = item.get("context", item.get("历史对话"))
            meta_val = item.get("metadata", item.get("元数据"))

            test_case = TestCase(
                input=str(input_val).strip(),
                actual_output=str(output_val).strip(),
                expected_output=str(item.get("expected_output", "")).strip() if item.get("expected_output") else None,
                context=_parse_string_to_list(context_val),
                retrieval_context=_parse_string_to_list(retrieval_val),  # 在这里调用解析函数
                metadata=_parse_string_to_dict(meta_val)  # 在这里解析字典
            )
            test_cases.append(test_case)

        return test_cases
